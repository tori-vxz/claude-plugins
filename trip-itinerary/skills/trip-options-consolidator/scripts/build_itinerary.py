#!/usr/bin/env python3
"""Build the trip-options-consolidator spreadsheet from merged trip data.

Usage:
    python3 build_itinerary.py <trip_data.json> [output.xlsx]

See ../templates/trip_data.example.json for the input schema, and
../references/spreadsheet-format.md for what the formatting below means
and why.
"""

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_FILL = PatternFill("solid", fgColor="404040")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=12)
TRAVEL_FILLS = [PatternFill("solid", fgColor="D9E1F2"), PatternFill("solid", fgColor="EAF0FA")]
TIP_FILL = PatternFill("solid", fgColor="33CC33")
ORDINARY_FILLS = [PatternFill("solid", fgColor="FFFFFF"), PatternFill("solid", fgColor="F2F2F2")]
NORMAL_FONT = Font(bold=False, size=11)

THIN = Side(style="thin")
ROW_BORDERS = {
    1: Border(left=THIN),
    2: Border(),
    3: Border(right=THIN),
    4: Border(),
    5: Border(left=THIN, right=THIN),
    6: Border(),
}

COLUMN_WIDTHS = {"A": 22, "B": 12, "C": 42, "D": 12, "E": 60, "F": 45}
HEADERS = ["Stop", "Type", "Item", "Budget Tier", "Details", "Link"]


def apply_row_borders(ws, row):
    for col, border in ROW_BORDERS.items():
        ws.cell(row=row, column=col).border = border


def write_header_row(ws):
    for col_idx, text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    apply_row_borders(ws, 1)


def write_section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = SECTION_FILL
    top_left = ws.cell(row=row, column=1, value=text)
    top_left.font = SECTION_FONT
    top_left.alignment = Alignment(vertical="center")
    apply_row_borders(ws, row)


def write_data_row(ws, row, stop, type_, item, tier, details, link, fill):
    for col_idx, value in enumerate([stop, type_, item, tier, details, link], start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 5))
    apply_row_borders(ws, row)


def write_travel_block(ws, row, label, options):
    write_section_header(ws, row, f"Travel: {label}")
    row += 1
    for i, opt in enumerate(options):
        fill = TRAVEL_FILLS[i % 2]
        write_data_row(ws, row, label, "Travel", opt["item"], opt["tier"], opt["details"], opt["link"], fill)
        row += 1
    return row


def write_city_block(ws, row, city, items):
    write_section_header(ws, row, city)
    row += 1
    ord_i = 0
    for item in items:
        if item["type"] == "Tip":
            fill = TIP_FILL
        else:
            fill = ORDINARY_FILLS[ord_i % 2]
            ord_i += 1
        write_data_row(ws, row, city, item["type"], item["item"], "", item["details"], item["link"], fill)
        row += 1
    return row


def build_sheet(ws, sequence, travel_legs, activities_for_tab):
    write_header_row(ws)
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    row = 2
    for step in sequence:
        if step["type"] == "leg":
            row = write_travel_block(ws, row, step["label"], travel_legs[step["label"]])
        elif step["type"] == "city":
            row = write_city_block(ws, row, step["name"], activities_for_tab[step["name"]])
        else:
            raise ValueError(f"Unknown sequence step type: {step['type']!r}")


def build_workbook(data):
    wb = Workbook()
    wb.remove(wb.active)
    sequence = data["sequence"]
    travel_legs = data["travel_legs"]
    for tab_name, activities_for_tab in data["activities"].items():
        ws = wb.create_sheet(tab_name)
        build_sheet(ws, sequence, travel_legs, activities_for_tab)
    return wb


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    data_path = Path(sys.argv[1])
    data = json.loads(data_path.read_text(encoding="utf-8"))

    if len(sys.argv) >= 3:
        out_path = Path(sys.argv[2])
    elif data.get("output_filename"):
        out_path = data_path.parent / data["output_filename"]
    else:
        cities = [s["name"] for s in data["sequence"] if s["type"] == "city"]
        out_path = data_path.parent / f"{'_'.join(cities)}_Trip_Itinerary.xlsx"

    wb = build_workbook(data)
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
