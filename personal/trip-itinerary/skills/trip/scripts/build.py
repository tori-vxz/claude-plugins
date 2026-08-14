#!/usr/bin/env python3
"""Build the trip itinerary spreadsheet from per-city shard files.

Usage:
    build.py <shard-dir> [--out-dir DIR] [--name PREFIX] [--format excel|artifact]

See scripts/USAGE.md for the shard schema, the gap-report semantics
(which gaps are fatal vs. warnings), and the full rendering-format
reference (colours, widths, borders, naming conventions).

Both formats are the same spreadsheet: same six columns, same tabs, same
section headers, same colors. `excel` (the default) writes an .xlsx
workbook; `artifact` writes a self-contained .html page to publish with
the Artifact tool. Only one is produced per run.
"""

import argparse
import html
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_BG = "404040"
HEADER_FG = "FFFFFF"
SECTION_BG = "1F4E78"
SECTION_FG = "FFFFFF"
TRAVEL_BGS = ["D9E1F2", "EAF0FA"]
TIP_BG = "33CC33"
ORDINARY_BGS = ["FFFFFF", "F2F2F2"]

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
HEADER_FONT = Font(color=HEADER_FG, bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor=SECTION_BG)
SECTION_FONT = Font(color=SECTION_FG, bold=True, size=12)
TRAVEL_FILLS = [PatternFill("solid", fgColor=c) for c in TRAVEL_BGS]
TIP_FILL = PatternFill("solid", fgColor=TIP_BG)
ORDINARY_FILLS = [PatternFill("solid", fgColor=c) for c in ORDINARY_BGS]
NORMAL_FONT = Font(bold=False, size=11)

THIN = Side(style="thin")
ROW_BORDERS = {
    1: Border(left=THIN),
    2: Border(),
    3: Border(right=THIN),
    4: Border(),
    5: Border(left=THIN, right=THIN),
    6: Border(),
}

COLUMN_WIDTHS = {"A": 22, "B": 12, "C": 42, "D": 20, "E": 60, "F": 45}
HEADERS = ["Stop", "Type", "Item", "Budget, Cost, Transfers", "Details", "Link"]

# Columns that wrap onto several lines inside one cell.
WRAPPED_COLUMNS = (4, 5)

HYPERLINK_FONT = Font(color="0563C1", underline="single", size=11)

ACTIVITY_LEVELS = ["Low Activity", "Medium Activity", "High Activity"]
TIERS = ["Cheapest", "Mid-range", "Luxury"]
TYPES = ["Activity", "Day trip", "Tip"]


# --- shard loading + validation ---------------------------------------------


def load_shards(shard_dir):
    """Glob, sort, and JSON-load shard files.

    Returns (shards, fatals, warnings) where shards is a list of
    (filename, number, raw_dict) tuples in trip order. Any shard that
    can't be parsed or lacks a usable 'city' is excluded from shards and
    recorded as fatal instead.
    """
    fatals = []
    warnings = []
    shard_files = sorted(shard_dir.glob("[0-9][0-9]-*.json"))
    if not shard_files:
        fatals.append(f"no shard files found in {shard_dir} matching NN-city.json")
        return [], fatals, warnings

    shards = []
    numbers = []
    for f in shard_files:
        try:
            num = int(f.name[:2])
        except ValueError:
            fatals.append(f"{f.name}: filename does not start with two digits")
            continue

        try:
            raw = json.loads(f.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            fatals.append(f"{f.name}: invalid JSON ({exc})")
            continue

        if not isinstance(raw, dict) or not isinstance(raw.get("city"), str) or not raw.get("city"):
            fatals.append(f"{f.name}: missing or invalid required 'city' field — shard cannot be placed in any tab")
            continue

        numbers.append(num)
        shards.append((f.name, num, raw))

    if numbers:
        if numbers[0] != 1:
            warnings.append(f"shard numbering does not start at 01 (starts at {numbers[0]:02d})")
        seen = set()
        dupes = set()
        for n in numbers:
            (dupes if n in seen else seen).add(n)
        if dupes:
            warnings.append(f"duplicate shard number(s): {', '.join(f'{n:02d}' for n in sorted(dupes))}")
        expected = set(range(numbers[0], max(numbers) + 1))
        missing = sorted(expected - set(numbers))
        if missing:
            warnings.append(f"sequence gap: missing shard number(s) {', '.join(f'{n:02d}' for n in missing)}")

    return shards, fatals, warnings


def normalize_travel_suboption(sub, context, warnings):
    """One entry of a tier's 'options' list — a single named service.

    A bare string is shorthand for {"details": sub}. Anything a suboption
    doesn't set falls back to the parent option's own item/cost/
    connections/link in expand_travel_option below.
    """
    if isinstance(sub, str):
        return {"details": sub}
    if not isinstance(sub, dict):
        warnings.append(f"{context}: an entry under 'options' is not an object or string — treated as empty")
        sub = {}
    link = sub.get("link")
    if isinstance(link, str) and link and not link.startswith("http"):
        warnings.append(f"{context}: sub-option link does not start with http: {link!r}")
    return {k: sub.get(k) for k in ("item", "cost", "connections", "details", "link") if isinstance(sub.get(k), str)}


def normalize_travel_option(opt, context, warnings):
    if not isinstance(opt, dict):
        warnings.append(f"{context}: travel option is not an object — treated as empty")
        opt = {}
    tier = opt.get("tier")
    if not isinstance(tier, str) or not tier:
        warnings.append(f"{context}: travel option missing 'tier'")
        tier = ""
    elif tier not in TIERS:
        warnings.append(f"{context}: unknown tier value {tier!r}")

    raw_sub_options = opt.get("options")
    sub_options = None
    if raw_sub_options is not None:
        if isinstance(raw_sub_options, list) and raw_sub_options:
            sub_options = [
                normalize_travel_suboption(s, f"{context} ({tier or 'unknown tier'})", warnings)
                for s in raw_sub_options
            ]
        else:
            warnings.append(
                f"{context} ({tier or 'unknown tier'}): 'options' present but not a non-empty list — ignored"
            )

    link = opt.get("link")
    if not isinstance(link, str):
        link = ""
    # A tier with its own 'options' list names a link per service instead;
    # only demand one here when there's no list to fall back on.
    if sub_options is None:
        if not link:
            warnings.append(f"{context} ({tier or 'unknown tier'}): missing link")
        elif not link.startswith("http"):
            warnings.append(f"{context} ({tier or 'unknown tier'}): link does not start with http: {link!r}")

    return {
        "tier": tier,
        "item": opt.get("item") if isinstance(opt.get("item"), str) else "",
        "cost": opt.get("cost") if isinstance(opt.get("cost"), str) else "",
        "connections": opt.get("connections") if isinstance(opt.get("connections"), str) else "",
        "details": opt.get("details") if isinstance(opt.get("details"), str) else "",
        "link": link,
        "options": sub_options,
    }


def normalize_leg(leg, context, warnings):
    """Validate + normalize an inbound_leg/trailing_leg dict.

    Returns (label, options) or (None, None) if the leg is unusable.
    """
    if not isinstance(leg, dict):
        warnings.append(f"{context}: leg is not an object — skipped")
        return None, None

    label = leg.get("label")
    if not isinstance(label, str) or not label:
        label = context
        warnings.append(f"{context}: leg missing 'label' — using fallback label {label!r}")

    raw_options = leg.get("options")
    if not isinstance(raw_options, list):
        warnings.append(f"{context}: leg missing 'options' list — treated as empty")
        raw_options = []

    if len(raw_options) != 3:
        warnings.append(f"{context}: leg has {len(raw_options)} option(s), expected exactly 3")

    options = [normalize_travel_option(o, context, warnings) for o in raw_options]
    tiers_seen = {o["tier"] for o in options}
    if tiers_seen != set(TIERS) and raw_options:
        warnings.append(
            f"{context}: leg tiers are {sorted(t for t in tiers_seen if t)}, expected exactly {TIERS}"
        )

    return label, options


def normalize_activity_item(item, context, warnings):
    if not isinstance(item, dict):
        warnings.append(f"{context}: activity item is not an object — treated as empty")
        item = {}

    type_ = item.get("type")
    if not isinstance(type_, str) or not type_:
        warnings.append(f"{context}: activity item missing 'type'")
        type_ = ""
    elif type_ not in TYPES:
        warnings.append(f"{context}: unknown type value {type_!r}")

    name = item.get("item") if isinstance(item.get("item"), str) else ""
    if not name:
        warnings.append(f"{context}: activity item missing 'item' (name)")

    link = item.get("link")
    if not isinstance(link, str) or not link:
        warnings.append(f"{context} ({name or 'unnamed'}): missing link")
        link = ""
    elif not link.startswith("http"):
        warnings.append(f"{context} ({name or 'unnamed'}): link does not start with http: {link!r}")

    return {
        "type": type_,
        "item": name,
        "details": item.get("details") if isinstance(item.get("details"), str) else "",
        "link": link,
    }


def assemble(shards, warnings):
    """Transpose city-major shards into the level-major shape the renderers
    expect: {sequence, travel_legs, activities}."""
    sequence = []
    travel_legs = {}
    activities = {level: {} for level in ACTIVITY_LEVELS}
    cities_order = []

    trailing_step = None
    num_shards = len(shards)
    seen_cities = set()

    for idx, (fname, num, raw) in enumerate(shards):
        city = raw["city"]
        if city in seen_cities:
            warnings.append(
                f"{fname}: duplicate city name {city!r} — its activities overwrite the earlier "
                f"shard's for {city!r}, and both sequence steps render the later shard's data"
            )
        seen_cities.add(city)
        cities_order.append(city)
        is_last = idx == num_shards - 1

        inbound = raw.get("inbound_leg")
        if inbound:
            label, options = normalize_leg(inbound, f"{fname} inbound_leg", warnings)
            if label is not None:
                if label in travel_legs:
                    warnings.append(f"{fname}: duplicate leg label {label!r} — overwriting previous entry")
                travel_legs[label] = options
                sequence.append({"type": "leg", "label": label})

        sequence.append({"type": "city", "name": city})

        trailing = raw.get("trailing_leg")
        if trailing:
            if is_last:
                label, options = normalize_leg(trailing, f"{fname} trailing_leg", warnings)
                if label is not None:
                    if label in travel_legs:
                        warnings.append(f"{fname}: duplicate leg label {label!r} — overwriting previous entry")
                    travel_legs[label] = options
                    trailing_step = {"type": "leg", "label": label}
            else:
                warnings.append(
                    f"{fname}: trailing_leg present on a non-final shard — ignored, "
                    "will not appear in the sequence"
                )

        raw_activities = raw.get("activities")
        if not isinstance(raw_activities, dict):
            warnings.append(f"{fname}: missing or invalid 'activities' object — all levels treated as empty")
            raw_activities = {}

        for level in ACTIVITY_LEVELS:
            items = raw_activities.get(level)
            if items is None:
                warnings.append(f"{fname}: activity level {level!r} missing for {city!r}")
                items = []
            elif not isinstance(items, list) or not items:
                warnings.append(f"{fname}: activity level {level!r} is empty for {city!r}")
                items = []
            activities[level][city] = [
                normalize_activity_item(it, f"{fname} {level}", warnings) for it in items
            ]

    if trailing_step is not None:
        sequence.append(trailing_step)

    return {"sequence": sequence, "travel_legs": travel_legs, "activities": activities}, cities_order


def print_gap_report(fatals, warnings):
    print("=== Gap report ===")
    if fatals:
        print(f"FATAL ({len(fatals)}):")
        for f in fatals:
            print(f"  - {f}")
    if warnings:
        print(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  - {w}")
    if not fatals and not warnings:
        print("No gaps found.")
    print("===================")


# --- row layout, shared by both renderers -----------------------------------


def split_details_text(text):
    parts = [text]
    for sep in ["\n", " | "]:
        expanded = []
        for part in parts:
            expanded.extend(part.split(sep))
        parts = expanded
    return [p.strip() for p in parts if p.strip()]


def tier_cell(tier, cost, connections):
    """The budget column: tier, then what it costs, then how many changes.

    Stacked on three lines inside the one cell, so a leg can be read down
    the column without crossing into Details. The tier itself is
    underlined and the two lines under it are not, which is why this comes
    back as a tuple for the two renderers to set rather than a plain
    string. Missing pieces are dropped rather than left as blank lines.
    """
    tier, cost, connections = (str(v).strip() if v else "" for v in (tier, cost, connections))
    if not (tier or cost or connections):
        return ""
    return ("budget", tier, cost, connections)


def expand_travel_option(opt):
    """One travel option in, one row per distinct service out.

    A tier that turned up several services keeps them apart: every row
    carries the same budget tier, and names the service it is actually
    for. See scripts/USAGE.md.
    """
    tier = opt.get("tier", "")
    cost = opt.get("cost", "")
    connections = opt.get("connections", "")
    parent_item = opt.get("item", "")
    parent_link = opt.get("link", "")
    details = opt.get("details", "")

    subs = opt.get("options")
    if subs is None and isinstance(details, list):
        subs = details
    if subs is None and isinstance(details, str):
        parts = split_details_text(details)
        if len(parts) > 1:
            subs = parts

    if not subs:
        return [
            {
                "item": parent_item,
                "tier": tier_cell(tier, cost, connections),
                "details": details if isinstance(details, str) else "",
                "link": parent_link,
            }
        ]

    rows = []
    for sub in subs:
        if isinstance(sub, str):
            rows.append(
                {
                    "item": parent_item,
                    "tier": tier_cell(tier, cost, connections),
                    "details": sub,
                    "link": parent_link,
                }
            )
        else:
            rows.append(
                {
                    "item": sub.get("item") or parent_item,
                    "tier": tier_cell(
                        tier,
                        sub.get("cost") or cost,
                        sub.get("connections") or connections,
                    ),
                    "details": sub.get("details", ""),
                    "link": sub.get("link") or parent_link,
                }
            )
    return rows


def expand_travel_options(options):
    rows = []
    for opt in options:
        rows.extend(expand_travel_option(opt))
    return rows


def sheet_rows(sequence, travel_legs, activities_for_tab):
    """Flatten one tab into a list of ('section'|'data', ...) rows.

    Both renderers below walk this, so the two formats cannot drift apart.
    """
    rows = []
    for step in sequence:
        if step["type"] == "leg":
            label = step["label"]
            rows.append(("section", f"Travel: {label}"))
            for i, opt in enumerate(expand_travel_options(travel_legs[label])):
                rows.append(
                    (
                        "data",
                        [label, "Travel", opt["item"], opt["tier"], opt["details"], opt["link"]],
                        "travel",
                        i % 2,
                    )
                )
        elif step["type"] == "city":
            city = step["name"]
            rows.append(("section", city))
            ord_i = 0
            for item in activities_for_tab[city]:
                if item["type"] == "Tip":
                    kind, stripe = "tip", 0
                else:
                    kind, stripe = "ordinary", ord_i % 2
                    ord_i += 1
                rows.append(
                    (
                        "data",
                        [city, item["type"], item["item"], "", item["details"], item["link"]],
                        kind,
                        stripe,
                    )
                )
        else:
            raise ValueError(f"Unknown sequence step type: {step['type']!r}")
    return rows


def tabs(data):
    sequence = data["sequence"]
    travel_legs = data["travel_legs"]
    return [
        (tab_name, sheet_rows(sequence, travel_legs, activities_for_tab))
        for tab_name, activities_for_tab in data["activities"].items()
    ]


# --- Excel --------------------------------------------------------------


def apply_row_borders(ws, row):
    for col, border in ROW_BORDERS.items():
        ws.cell(row=row, column=col).border = border


def write_header_row(ws):
    for col_idx, text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    apply_row_borders(ws, 1)


def write_section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    top_left = ws.cell(row=row, column=1, value=text)
    top_left.font = SECTION_FONT
    top_left.alignment = Alignment(vertical="center")
    apply_row_borders(ws, row)


def budget_rich_text(tier, cost, connections):
    """The budget cell as Excel rich text: tier underlined, rest plain."""
    runs = []
    if tier:
        runs.append(TextBlock(InlineFont(u="single", sz=11), tier))
    tail = "\n".join(v for v in (cost, connections) if v)
    if tail:
        runs.append(("\n" + tail) if runs else tail)
    return CellRichText(runs)


def write_data_row(ws, row, values, fill):
    for col_idx, value in enumerate(values, start=1):
        if isinstance(value, tuple) and value[0] == "budget":
            value = budget_rich_text(*value[1:])
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in WRAPPED_COLUMNS))
    apply_row_borders(ws, row)
    # Real openpyxl hyperlink on column F, not a bare string. Set link +
    # font by hand (rather than the "Hyperlink" named style) so the row's
    # fill/border set above survive untouched.
    link_cell = ws.cell(row=row, column=6)
    link_value = values[5]
    if isinstance(link_value, str) and link_value.startswith("http"):
        link_cell.hyperlink = link_value
        link_cell.font = HYPERLINK_FONT


def build_sheet(ws, rows):
    write_header_row(ws)
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    row = 2
    for entry in rows:
        if entry[0] == "section":
            write_section_header(ws, row, entry[1])
        else:
            _, values, kind, stripe = entry
            if kind == "travel":
                fill = TRAVEL_FILLS[stripe]
            elif kind == "tip":
                fill = TIP_FILL
            else:
                fill = ORDINARY_FILLS[stripe]
            write_data_row(ws, row, values, fill)
        row += 1


def build_workbook(data):
    wb = Workbook()
    wb.remove(wb.active)
    for tab_name, rows in tabs(data):
        build_sheet(wb.create_sheet(tab_name), rows)
    return wb


# --- Artifact -------------------------------------------------------------

PAGE_CSS = """
:root {{
  --page-bg: #FFFFFF;
  --ink: #1a1a1a;
  --rule: #c8c8c8;
}}
body {{
  background: var(--page-bg);
  color: var(--ink);
  font-family: Calibri, Carlito, "Segoe UI", system-ui, sans-serif;
  font-size: 14px;
  margin: 0;
  padding: 24px;
}}
h1 {{ font-size: 20px; margin: 0 0 16px; }}
.tabs {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 0; }}
.tabs button {{
  font: inherit;
  font-weight: 700;
  padding: 8px 16px;
  border: 1px solid var(--rule);
  border-bottom: none;
  border-radius: 4px 4px 0 0;
  background: #{ordinary1};
  color: var(--ink);
  cursor: pointer;
}}
.tabs button[aria-selected="true"] {{ background: #{section_bg}; color: #{section_fg}; }}
.sheet {{ overflow-x: auto; border: 1px solid var(--rule); }}
table {{ border-collapse: collapse; width: 100%; min-width: 1100px; }}
th, td {{
  text-align: left;
  vertical-align: top;
  padding: 4px 6px;
  border: none;
}}
th {{ background: #{header_bg}; color: #{header_fg}; font-weight: 700; }}
tr.section td {{
  background: #{section_bg};
  color: #{section_fg};
  font-weight: 700;
  font-size: 15px;
  vertical-align: middle;
}}
td.c1, td.c5, th.c1, th.c5 {{ border-left: 1px solid var(--ink); }}
td.c3, td.c5, th.c3, th.c5 {{ border-right: 1px solid var(--ink); }}
col.c1 {{ width: 150px; }}
col.c2 {{ width: 82px; }}
col.c3 {{ width: 285px; }}
col.c4 {{ width: 136px; }}
col.c5 {{ width: 410px; }}
col.c6 {{ width: 305px; }}
td.c4 u {{ text-underline-offset: 2px; }}
a {{ color: #1155cc; word-break: break-all; }}
[hidden] {{ display: none; }}
"""


def cell_html(value, col):
    if isinstance(value, tuple) and value[0] == "budget":
        _, tier, cost, connections = value
        lines = []
        if tier:
            lines.append(f"<u>{html.escape(tier)}</u>")
        lines.extend(html.escape(v) for v in (cost, connections) if v)
        return "<br>".join(lines)
    text = html.escape(str(value or ""))
    if col == 6 and value:
        return f'<a href="{html.escape(str(value))}" target="_blank" rel="noopener">{text}</a>'
    return text


def build_html(data, title):
    all_tabs = tabs(data)
    parts = [f"<title>{html.escape(title)}</title>"]
    parts.append(
        "<style>"
        + PAGE_CSS.format(
            header_bg=HEADER_BG,
            header_fg=HEADER_FG,
            section_bg=SECTION_BG,
            section_fg=SECTION_FG,
            ordinary1=ORDINARY_BGS[1],
        )
        + "</style>"
    )
    parts.append(f"<h1>{html.escape(title)}</h1>")

    parts.append('<div class="tabs" role="tablist">')
    for i, (tab_name, _) in enumerate(all_tabs):
        selected = "true" if i == 0 else "false"
        parts.append(
            f'<button role="tab" aria-selected="{selected}" aria-controls="sheet{i}" '
            f'id="tab{i}">{html.escape(tab_name)}</button>'
        )
    parts.append("</div>")

    for i, (tab_name, rows) in enumerate(all_tabs):
        hidden = "" if i == 0 else " hidden"
        parts.append(
            f'<div class="sheet" role="tabpanel" id="sheet{i}" aria-labelledby="tab{i}"{hidden}>'
        )
        parts.append("<table>")
        parts.append("<colgroup>" + "".join(f'<col class="c{c}">' for c in range(1, 7)) + "</colgroup>")
        parts.append(
            "<thead><tr>"
            + "".join(f'<th class="c{c}">{html.escape(h)}</th>' for c, h in enumerate(HEADERS, start=1))
            + "</tr></thead><tbody>"
        )
        for entry in rows:
            if entry[0] == "section":
                parts.append(
                    f'<tr class="section"><td class="c1" colspan="6">{html.escape(entry[1])}</td></tr>'
                )
            else:
                _, values, kind, stripe = entry
                if kind == "travel":
                    bg = TRAVEL_BGS[stripe]
                elif kind == "tip":
                    bg = TIP_BG
                else:
                    bg = ORDINARY_BGS[stripe]
                cells = "".join(
                    f'<td class="c{c}">{cell_html(v, c)}</td>' for c, v in enumerate(values, start=1)
                )
                parts.append(f'<tr style="background:#{bg}">{cells}</tr>')
        parts.append("</tbody></table></div>")

    parts.append(
        "<script>\n"
        "const tabs = [...document.querySelectorAll('[role=tab]')];\n"
        "tabs.forEach((tab, i) => tab.addEventListener('click', () => {\n"
        "  tabs.forEach((t, j) => {\n"
        "    t.setAttribute('aria-selected', String(i === j));\n"
        "    document.getElementById('sheet' + j).hidden = i !== j;\n"
        "  });\n"
        "}));\n"
        "</script>"
    )
    return "\n".join(parts)


def main():
    parser = argparse.ArgumentParser(
        description="Build the trip itinerary spreadsheet from per-city shard files.",
    )
    parser.add_argument("shard_dir", type=Path, help="Directory containing NN-city.json shard files")
    parser.add_argument("--out-dir", type=Path, default=None, help="Output directory (default: shard-dir)")
    parser.add_argument(
        "--name",
        default=None,
        help="Filename prefix (default: derived from city names, e.g. Busan_Seoul_Incheon)",
    )
    parser.add_argument(
        "--format",
        choices=["excel", "artifact"],
        default="excel",
        help="excel (default) writes an .xlsx workbook; artifact writes a self-contained .html page",
    )
    args = parser.parse_args()

    shard_dir = args.shard_dir
    if not shard_dir.is_dir():
        print(f"FATAL: {shard_dir} is not a directory")
        sys.exit(1)

    shards, fatals, warnings = load_shards(shard_dir)

    data = None
    cities_order = []
    if not fatals:
        data, cities_order = assemble(shards, warnings)

    print_gap_report(fatals, warnings)

    if fatals:
        sys.exit(1)

    out_dir = args.out_dir or shard_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.name:
        stem = f"{args.name}_Trip_Itinerary"
    elif cities_order:
        stem = f"{'_'.join(cities_order)}_Trip_Itinerary"
    else:
        stem = "Trip_Itinerary"

    if args.format == "excel":
        out_path = out_dir / f"{stem}.xlsx"
        build_workbook(data).save(out_path)
    else:
        out_path = out_dir / f"{stem}.html"
        title = stem.replace("_", " ")
        out_path.write_text(build_html(data, title), encoding="utf-8")

    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
