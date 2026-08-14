#!/usr/bin/env python3
"""Build the trip-options-consolidator spreadsheet from merged trip data.

Usage:
    python3 build_itinerary.py <trip_data.json> [output] [--format excel|artifact]

Both formats are the same spreadsheet: same six columns, same tabs, same
section headers, same colors. `excel` writes an .xlsx workbook; `artifact`
writes a self-contained .html page to publish with the Artifact tool.

See ../templates/trip_data.example.json for the input schema, and
../references/spreadsheet-format.md for what the formatting below means
and why.
"""

import html
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HEADER_BG = "404040"
HEADER_FG = "FFFFFF"
SECTION_BG = "1F4E78"
SECTION_FG = "FFFFFF"
TRAVEL_BGS = ["D9E1F2", "EAF0FA"]
TIP_BG = "33CC33"
ORDINARY_BGS = ["FFFFFF", "F2F2F2"]

HEADER_FILL = PatternFill("solid", fgColor=HEADER_BG)
HEADER_FONT = Font(color=HEADER_FG, bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor=SECTION_BG)
SECTION_FONT = Font(color=SECTION_FG, bold=True, size=12)
TRAVEL_FILLS = [PatternFill("solid", fgColor=c) for c in TRAVEL_BGS]
TIP_FILL = PatternFill("solid", fgColor=TIP_BG)
ORDINARY_FILLS = [PatternFill("solid", fgColor=c) for c in ORDINARY_BGS]
NORMAL_FONT = Font(bold=False, size=11)

THIN = Side(style="thin")
ROW_BORDERS = {
    1: Border(left=THIN),
    2: Border(),
    3: Border(right=THIN),
    4: Border(),
    5: Border(left=THIN, right=THIN),
    6: Border(),
}

COLUMN_WIDTHS = {"A": 22, "B": 12, "C": 42, "D": 20, "E": 60, "F": 45}
HEADERS = ["Stop", "Type", "Item", "Budget, Cost, Transfers", "Details", "Link"]

# Columns that wrap onto several lines inside one cell.
WRAPPED_COLUMNS = (4, 5)

# Details text carrying more than one travel option is split on these, so a
# leg that came back packed into one cell still lands one option per row.
DETAIL_SEPARATORS = ["\n", " | "]


def split_details_text(text):
    parts = [text]
    for sep in DETAIL_SEPARATORS:
        expanded = []
        for part in parts:
            expanded.extend(part.split(sep))
        parts = expanded
    return [p.strip() for p in parts if p.strip()]


def tier_cell(tier, cost, connections):
    """The budget column: tier, then what it costs, then how many changes.

    Stacked on three lines inside the one cell, so a leg can be read down
    the column without crossing into Details. The tier itself is
    underlined and the two lines under it are not, which is why this comes
    back as a tuple for the two builders to set rather than a plain
    string. Missing pieces are dropped rather than left as blank lines.
    """
    tier, cost, connections = (str(v).strip() if v else "" for v in (tier, cost, connections))
    if not (tier or cost or connections):
        return ""
    return ("budget", tier, cost, connections)


def expand_travel_option(opt):
    """One travel option in, one row per distinct service out.

    A tier that turned up several services keeps them apart: every row
    carries the same budget tier, and names the service it is actually
    for. See references/spreadsheet-format.md.
    """
    tier = opt.get("tier", "")
    cost = opt.get("cost", "")
    connections = opt.get("connections", "")
    parent_item = opt.get("item", "")
    parent_link = opt.get("link", "")
    details = opt.get("details", "")

    subs = opt.get("options")
    if subs is None and isinstance(details, list):
        subs = details
    if subs is None and isinstance(details, str):
        parts = split_details_text(details)
        if len(parts) > 1:
            subs = parts

    if not subs:
        return [
            {
                "item": parent_item,
                "tier": tier_cell(tier, cost, connections),
                "details": details if isinstance(details, str) else "",
                "link": parent_link,
            }
        ]

    rows = []
    for sub in subs:
        if isinstance(sub, str):
            rows.append(
                {
                    "item": parent_item,
                    "tier": tier_cell(tier, cost, connections),
                    "details": sub,
                    "link": parent_link,
                }
            )
        else:
            rows.append(
                {
                    "item": sub.get("item") or parent_item,
                    "tier": tier_cell(
                        tier,
                        sub.get("cost") or cost,
                        sub.get("connections") or connections,
                    ),
                    "details": sub.get("details", ""),
                    "link": sub.get("link") or parent_link,
                }
            )
    return rows


def expand_travel_options(options):
    rows = []
    for opt in options:
        rows.extend(expand_travel_option(opt))
    return rows


def sheet_rows(sequence, travel_legs, activities_for_tab):
    """Flatten one tab into a list of ('section'|'data', ...) rows.

    Both builders below walk this, so the two formats cannot drift apart.
    """
    rows = []
    for step in sequence:
        if step["type"] == "leg":
            label = step["label"]
            rows.append(("section", f"Travel: {label}"))
            for i, opt in enumerate(expand_travel_options(travel_legs[label])):
                rows.append(
                    (
                        "data",
                        [label, "Travel", opt["item"], opt["tier"], opt["details"], opt["link"]],
                        "travel",
                        i % 2,
                    )
                )
        elif step["type"] == "city":
            city = step["name"]
            rows.append(("section", city))
            ord_i = 0
            for item in activities_for_tab[city]:
                if item["type"] == "Tip":
                    kind, stripe = "tip", 0
                else:
                    kind, stripe = "ordinary", ord_i % 2
                    ord_i += 1
                rows.append(
                    (
                        "data",
                        [city, item["type"], item["item"], "", item["details"], item["link"]],
                        kind,
                        stripe,
                    )
                )
        else:
            raise ValueError(f"Unknown sequence step type: {step['type']!r}")
    return rows


# --- Excel ------------------------------------------------------------------


def apply_row_borders(ws, row):
    for col, border in ROW_BORDERS.items():
        ws.cell(row=row, column=col).border = border


def write_header_row(ws):
    for col_idx, text in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    apply_row_borders(ws, 1)


def write_section_header(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    top_left = ws.cell(row=row, column=1, value=text)
    top_left.font = SECTION_FONT
    top_left.alignment = Alignment(vertical="center")
    apply_row_borders(ws, row)


def budget_rich_text(tier, cost, connections):
    """The budget cell as Excel rich text: tier underlined, rest plain."""
    runs = []
    if tier:
        runs.append(TextBlock(InlineFont(u="single", sz=11), tier))
    tail = "\n".join(v for v in (cost, connections) if v)
    if tail:
        runs.append(("\n" + tail) if runs else tail)
    return CellRichText(runs)


def write_data_row(ws, row, values, fill):
    for col_idx, value in enumerate(values, start=1):
        if isinstance(value, tuple) and value[0] == "budget":
            value = budget_rich_text(*value[1:])
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = NORMAL_FONT
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in WRAPPED_COLUMNS))
    apply_row_borders(ws, row)


def build_sheet(ws, rows):
    write_header_row(ws)
    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width
    row = 2
    for entry in rows:
        if entry[0] == "section":
            write_section_header(ws, row, entry[1])
        else:
            _, values, kind, stripe = entry
            if kind == "travel":
                fill = TRAVEL_FILLS[stripe]
            elif kind == "tip":
                fill = TIP_FILL
            else:
                fill = ORDINARY_FILLS[stripe]
            write_data_row(ws, row, values, fill)
        row += 1


def build_workbook(data):
    wb = Workbook()
    wb.remove(wb.active)
    for tab_name, rows in tabs(data):
        build_sheet(wb.create_sheet(tab_name), rows)
    return wb


# --- Artifact ---------------------------------------------------------------

PAGE_CSS = """
:root {{
  --page-bg: #FFFFFF;
  --ink: #1a1a1a;
  --rule: #c8c8c8;
}}
body {{
  background: var(--page-bg);
  color: var(--ink);
  font-family: Calibri, Carlito, "Segoe UI", system-ui, sans-serif;
  font-size: 14px;
  margin: 0;
  padding: 24px;
}}
h1 {{ font-size: 20px; margin: 0 0 16px; }}
.tabs {{ display: flex; flex-wrap: wrap; gap: 4px; margin-bottom: 0; }}
.tabs button {{
  font: inherit;
  font-weight: 700;
  padding: 8px 16px;
  border: 1px solid var(--rule);
  border-bottom: none;
  border-radius: 4px 4px 0 0;
  background: #{ordinary1};
  color: var(--ink);
  cursor: pointer;
}}
.tabs button[aria-selected="true"] {{ background: #{section_bg}; color: #{section_fg}; }}
.sheet {{ overflow-x: auto; border: 1px solid var(--rule); }}
table {{ border-collapse: collapse; width: 100%; min-width: 1100px; }}
th, td {{
  text-align: left;
  vertical-align: top;
  padding: 4px 6px;
  border: none;
}}
th {{ background: #{header_bg}; color: #{header_fg}; font-weight: 700; }}
tr.section td {{
  background: #{section_bg};
  color: #{section_fg};
  font-weight: 700;
  font-size: 15px;
  vertical-align: middle;
}}
td.c1, td.c5, th.c1, th.c5 {{ border-left: 1px solid var(--ink); }}
td.c3, td.c5, th.c3, th.c5 {{ border-right: 1px solid var(--ink); }}
col.c1 {{ width: 150px; }}
col.c2 {{ width: 82px; }}
col.c3 {{ width: 285px; }}
col.c4 {{ width: 136px; }}
col.c5 {{ width: 410px; }}
col.c6 {{ width: 305px; }}
td.c4 u {{ text-underline-offset: 2px; }}
a {{ color: #1155cc; word-break: break-all; }}
[hidden] {{ display: none; }}
"""


def cell_html(value, col):
    if isinstance(value, tuple) and value[0] == "budget":
        _, tier, cost, connections = value
        lines = []
        if tier:
            lines.append(f"<u>{html.escape(tier)}</u>")
        lines.extend(html.escape(v) for v in (cost, connections) if v)
        return "<br>".join(lines)
    text = html.escape(str(value or ""))
    if col == 6 and value:
        return f'<a href="{html.escape(str(value))}" target="_blank" rel="noopener">{text}</a>'
    return text


def build_html(data, title):
    all_tabs = tabs(data)
    parts = [f"<title>{html.escape(title)}</title>"]
    parts.append(
        "<style>"
        + PAGE_CSS.format(
            header_bg=HEADER_BG,
            header_fg=HEADER_FG,
            section_bg=SECTION_BG,
            section_fg=SECTION_FG,
            ordinary1=ORDINARY_BGS[1],
        )
        + "</style>"
    )
    parts.append(f"<h1>{html.escape(title)}</h1>")

    parts.append('<div class="tabs" role="tablist">')
    for i, (tab_name, _) in enumerate(all_tabs):
        selected = "true" if i == 0 else "false"
        parts.append(
            f'<button role="tab" aria-selected="{selected}" aria-controls="sheet{i}" '
            f'id="tab{i}">{html.escape(tab_name)}</button>'
        )
    parts.append("</div>")

    for i, (tab_name, rows) in enumerate(all_tabs):
        hidden = "" if i == 0 else " hidden"
        parts.append(
            f'<div class="sheet" role="tabpanel" id="sheet{i}" aria-labelledby="tab{i}"{hidden}>'
        )
        parts.append("<table>")
        parts.append("<colgroup>" + "".join(f'<col class="c{c}">' for c in range(1, 7)) + "</colgroup>")
        parts.append(
            "<thead><tr>"
            + "".join(f'<th class="c{c}">{html.escape(h)}</th>' for c, h in enumerate(HEADERS, start=1))
            + "</tr></thead><tbody>"
        )
        for entry in rows:
            if entry[0] == "section":
                parts.append(
                    f'<tr class="section"><td class="c1" colspan="6">{html.escape(entry[1])}</td></tr>'
                )
            else:
                _, values, kind, stripe = entry
                if kind == "travel":
                    bg = TRAVEL_BGS[stripe]
                elif kind == "tip":
                    bg = TIP_BG
                else:
                    bg = ORDINARY_BGS[stripe]
                cells = "".join(
                    f'<td class="c{c}">{cell_html(v, c)}</td>' for c, v in enumerate(values, start=1)
                )
                parts.append(f'<tr style="background:#{bg}">{cells}</tr>')
        parts.append("</tbody></table></div>")

    parts.append(
        "<script>\n"
        "const tabs = [...document.querySelectorAll('[role=tab]')];\n"
        "tabs.forEach((tab, i) => tab.addEventListener('click', () => {\n"
        "  tabs.forEach((t, j) => {\n"
        "    t.setAttribute('aria-selected', String(i === j));\n"
        "    document.getElementById('sheet' + j).hidden = i !== j;\n"
        "  });\n"
        "}));\n"
        "</script>"
    )
    return "\n".join(parts)


# --- Shared -----------------------------------------------------------------


def tabs(data):
    sequence = data["sequence"]
    travel_legs = data["travel_legs"]
    return [
        (tab_name, sheet_rows(sequence, travel_legs, activities_for_tab))
        for tab_name, activities_for_tab in data["activities"].items()
    ]


def default_stem(data):
    if data.get("output_filename"):
        return Path(data["output_filename"]).stem
    cities = [s["name"] for s in data["sequence"] if s["type"] == "city"]
    return f"{'_'.join(cities)}_Trip_Itinerary"


def main():
    args = [a for a in sys.argv[1:]]
    fmt = "excel"
    if "--format" in args:
        i = args.index("--format")
        fmt = args[i + 1]
        del args[i : i + 2]

    if not args or fmt not in ("excel", "artifact"):
        print(__doc__)
        sys.exit(1)

    data_path = Path(args[0])
    data = json.loads(data_path.read_text(encoding="utf-8"))

    stem = default_stem(data)
    suffix = ".xlsx" if fmt == "excel" else ".html"
    out_path = Path(args[1]) if len(args) >= 2 else data_path.parent / f"{stem}{suffix}"

    if fmt == "excel":
        build_workbook(data).save(out_path)
    else:
        title = stem.replace("_", " ")
        out_path.write_text(build_html(data, title), encoding="utf-8")

    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
